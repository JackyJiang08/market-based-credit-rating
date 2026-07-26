"""Submission workbook: the deliverable.

Writes one timestamped workbook per run (never overwriting -- an existing path
raises rather than being replaced) with four sheets:
  - `README`     : provenance (model version, git SHA, data vintage) and every
                   convention in force, with its sweep range where one exists.
  - `Ratings`    : the presentation rule -- RiskScore first, letter only with
                   its interval attached.
  - `Asset`      : one row per company, exactly `records.ASSET_SCHEMA` in order.
  - `validation` : diagnostics per company -- data status, EM convergence, drift
                   regime/t, rating basis, determination, floor/scale-top flags,
                   bootstrap interval, convention span, field provenance.

Filename stamps follow the team timestamp standard (docs/TIMING_PROTOCOL.md
§10): tz-aware UTC, compact ISO 8601.

This module owns the file; it does not own the schema. Every field comes from
`dashboard.records`, so the deliverable, the per-company workbooks and the long
table cannot drift apart. Outputs are generated artifacts and are git-ignored.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime, timezone

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from creditrating._paths import REPO_ROOT as _PROJECT_ROOT

from ..data.company import CompanyData
from . import records

LOG = logging.getLogger(__name__)


OUTPUT_DIR = os.path.join(_PROJECT_ROOT, "outputs")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _format(ws) -> None:
    ws.freeze_panes = "B2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((len(str(x.value)) for x in col if x.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), 40)


def write_submission(companies: list[CompanyData], filename: str | None = None) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if filename is None:
        # Team timestamp standard (docs/TIMING_PROTOCOL.md §10): UTC, compact
        # ISO 8601 -- never naive local machine time.
        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        filename = f"submission_{stamp}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    if os.path.exists(path):
        raise FileExistsError(
            f"{path} already exists; a submission is never overwritten. "
            "Re-run to get a fresh stamp, or pass a different filename."
        )

    # All frames come from io.records, the single source of truth for
    # every published credit field. The Asset sheet is guaranteed to carry
    # exactly records.ASSET_SCHEMA, in order.
    readme = records.readme_frame(companies)
    ratings = records.ratings_frame(companies)
    asset = records.asset_frame(companies)
    validation = records.validation_frame(companies)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # README first: a reader opening the file lands on the provenance and
        # the conventions before the numbers. Ratings second: the presentation
        # rule (RiskScore first, letter with interval) before the canonical form.
        readme.to_excel(writer, sheet_name="README", index=False)
        ratings.to_excel(writer, sheet_name="Ratings", index=False)
        asset.to_excel(writer, sheet_name="Asset", index=False)
        validation.to_excel(writer, sheet_name="validation", index=False)
        _format(writer.book["README"])
        _format(writer.book["Ratings"])
        _format(writer.book["Asset"])
        _format(writer.book["validation"])
        # The README is prose; give it room rather than content-fitting.
        ws = writer.book["README"]
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 110
        for row in ws.iter_rows(min_row=2):
            # A fresh Alignment, not .copy(**kw) on the existing one -- the
            # kwargs form of StyleProxy.copy is deprecated in openpyxl 3.1.
            row[1].alignment = Alignment(wrap_text=True, vertical="top")

    LOG.info("submission workbook -> %s", os.path.relpath(path, _PROJECT_ROOT))
    return path
