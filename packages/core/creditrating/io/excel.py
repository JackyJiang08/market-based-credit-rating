"""Output layer: formatted Excel workbooks and the tidy long table.

Three deliverables, each for a different consumer:
  - per-company workbook  -> a human analyst drilling into one name
  - master workbook       -> quick cross-company comparison
  - tidy long table       -> databases / BI / pivot analysis
"""

from __future__ import annotations

import logging
import os
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..data.company import CompanyData
from ..data import provider_config as raw_config
from ..data import provenance as lineage
from ..model import config as signal_config

from . import config, records

LOG = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


# --------------------------------------------------------------------------- #
# Formatting helpers
# --------------------------------------------------------------------------- #
def _format_sheet(worksheet, *, freeze: str = "B2") -> None:
    """Header styling, frozen panes, and content-based column widths."""
    worksheet.freeze_panes = freeze
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col_cells in worksheet.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None),
                      default=10)
        worksheet.column_dimensions[letter].width = min(max(longest + 2, 12), 48)


def fmt_market_cap(value: Optional[float]) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "N/A"
    for unit, size in (("T", 1e12), ("B", 1e9), ("M", 1e6), ("K", 1e3)):
        if abs(value) >= size:
            return f"{value / size:.2f}{unit}"
    return f"{value:.0f}"


def _pct(x: Optional[float]) -> str:
    return f"{x*100:.3f}%" if x is not None and pd.notna(x) else "N/A"


# --------------------------------------------------------------------------- #
# Per-company workbook
# --------------------------------------------------------------------------- #
def write_company_workbook(data: CompanyData, years: int) -> str:
    path = os.path.join(config.OUTPUT_DIR, f"{data.ticker}_data.xlsx")

    summary = pd.DataFrame({
        "Field": [
            "Ticker", "Company", "Sector", "Industry", "Currency",
            "Market Cap", "Market Cap (raw)",
            "Shares Outstanding (traded class)",
            "Shares Outstanding (reference, mktcap/price)",
            "Last Closing Price", "Dividend Rate (annual)", "Dividend Yield",
            "Financials Window",
            "Extracted (as-of)", "Equity Source", "Rates Source",
        ],
        "Value": [
            data.ticker, data.name, data.sector, data.industry, data.currency,
            fmt_market_cap(data.market_cap), data.market_cap,
            data.shares_traded_class, data.reference_shares,
            data.last_close, data.dividend_rate,
            f"{data.dividend_yield:.2f}%" if data.dividend_yield else "N/A",
            f"{years}y",
            data.as_of, lineage.EQUITY_SOURCE, lineage.RATES_SOURCE,
        ],
    })

    rating_df = pd.DataFrame()
    if data.sigma_A is not None:
        # Projected from the shared record (dashboard.records) so this sheet
        # cannot disagree with the deliverable Asset sheet.
        r = records.credit_record(data)
        rating_df = pd.DataFrame({
            "Field": [
                "Equity E (latest)", "Default-point Debt D", "Risk-free r (1Y)",
                "Horizon T (years)", "Asset Volatility sigma_A", "Asset Value A",
                "Asset Return eta_A", "R (eta - sigma^2/2)", "CCM",
                "mu (life expectancy)", "lambda (default peak)", "TiC",
                "TiC Risk Score", "Distance to Default", "EDF", "PIT PD",
                "TTC PD", "S&P Rating", "Rating Basis", "Outlook (PIT - TTC)",
                "Drift regime", "Drift SE", "Drift span (years)",
                "EM iters / converged",
            ],
            "Value": [
                r["equity"], r["default_point_debt"], _pct(r["interest_rate"]),
                signal_config.HORIZON_YEARS, _pct(r["sigma_A"]), r["asset_value"],
                _pct(r["eta_A"]), _pct(r["drift"]), _round(r["ccm"]),
                _round(r["mu"], 2), _round(r["lam"], 4), _round(r["tic"]),
                _round(r["risk_score"], 2), _round(r["dd"], 2), _pct(r["edf"]),
                _pct(r["pit_pd"]), _pct(r["ttc_pd"]), r["sp_rating"] or "N/A",
                r["rating_basis"] or "N/A", _pct(r["outlook"]),
                r["drift_regime"] or "N/A", _pct(r["drift_se"]),
                _round(r["drift_span_years"], 2),
                f"{r['em_iters']} / {r['em_converged']}",
            ],
        })

    sheets: list[tuple[str, pd.DataFrame, bool]] = [
        ("Summary", summary, False),
        ("Credit Rating", rating_df, False),
        ("Aligned Panel", data.panel, True),
        ("Debt & Liabilities", data.debt_schedule, True),
        ("Price History", data.prices, True),
        ("Dividends", data.dividends, True),
        ("Q Income Statement", data.q_income, True),
        ("Q Balance Sheet", data.q_balance, True),
        ("Q Cash Flow", data.q_cashflow, True),
        ("Annual Income Statement", data.a_income, True),
        ("Annual Balance Sheet", data.a_balance, True),
        ("Annual Cash Flow", data.a_cashflow, True),
    ]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df, with_index in sheets:
            if df is None or df.empty:
                continue
            df.to_excel(writer, sheet_name=name[:31], index=with_index)
        for name, df, _ in sheets:
            if df is None or df.empty:
                continue
            _format_sheet(writer.book[name[:31]])
    LOG.info("  -> saved %s", os.path.basename(path))
    return path


def _panel_last(data: CompanyData, col: str) -> Optional[float]:
    if data.panel is None or data.panel.empty or col not in data.panel:
        return None
    s = data.panel[col].dropna()
    return None if s.empty else float(s.iloc[-1])


def _round(x: Optional[float], nd: int = 4) -> Optional[float]:
    return round(x, nd) if x is not None and pd.notna(x) else None


# --------------------------------------------------------------------------- #
# Master workbook
# --------------------------------------------------------------------------- #
def write_master_workbook(companies: list[CompanyData], rates: pd.DataFrame) -> str:
    path = os.path.join(config.OUTPUT_DIR, "_MASTER_summary.xlsx")

    summary = pd.DataFrame([{
        "Ticker": c.ticker,
        "Company": c.name,
        "Sector": c.sector,
        "Market Cap": fmt_market_cap(c.market_cap),
        "Market Cap (raw)": c.market_cap,
        "Reference Shares": c.reference_shares,
        "Last Close": c.last_close,
        "Dividend Rate": c.dividend_rate,
        "Dividend Yield (%)": c.dividend_yield,
    } for c in companies])

    debt_rows = []
    for c in companies:
        if c.debt_schedule.empty:
            continue
        latest = c.debt_schedule.columns[0]
        row = {"Ticker": c.ticker, "Period": latest}
        row.update(c.debt_schedule[latest].to_dict())
        debt_rows.append(row)
    debt_latest = pd.DataFrame(debt_rows)

    rating_rows = []
    for c in companies:
        if c.sigma_A is None:
            continue
        r = records.credit_record(c)
        rating_rows.append({
            "Ticker": r["symbol"],
            "sigma_A": r["sigma_A"],
            "eta_A": r["eta_A"],
            "AssetValue_A": r["asset_value"],
            "DefaultPointDebt_D": r["default_point_debt"],
            "CCM": r["ccm"],
            "mu": r["mu"],
            "lambda": r["lam"],
            "TiC_Risk_Score": r["risk_score"],
            "DD": r["dd"],
            "EDF": r["edf"],
            "PIT_PD": r["pit_pd"],
            "TTC_PD": r["ttc_pd"],
            "SP_Rating": r["sp_rating"],
            "Rating_Basis": r["rating_basis"],
            "Outlook": r["outlook"],
            "Drift_Regime": r["drift_regime"],
        })
    ratings_df = pd.DataFrame(rating_rows)

    rates_display = pd.DataFrame()
    if rates is not None and not rates.empty:
        rates_display = rates.rename(columns=raw_config.FRED_SERIES)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Company Summary", index=False)
        _format_sheet(writer.book["Company Summary"])
        if not ratings_df.empty:
            ratings_df.to_excel(writer, sheet_name="Ratings", index=False)
            _format_sheet(writer.book["Ratings"])
        if not debt_latest.empty:
            debt_latest.to_excel(writer, sheet_name="Debt & Liab (latest)", index=False)
            _format_sheet(writer.book["Debt & Liab (latest)"])
        if not rates_display.empty:
            rates_display.to_excel(writer, sheet_name="Macro Rates", index=False)
            _format_sheet(writer.book["Macro Rates"])
    LOG.info("master workbook -> %s", os.path.basename(path))
    return path
