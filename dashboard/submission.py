"""Submission workbook: the deliverable.

Writes one timestamped workbook per run (never overwriting a prior submission)
with two sheets:
  - `Asset`      : one row per company, exactly `records.ASSET_SCHEMA` in order.
  - `validation` : diagnostics per company -- data status, EM convergence, drift
                   regime and standard error, rating basis, floor flag.

This module owns the file; it does not own the schema. Every field comes from
`dashboard.records`, so the deliverable, the per-company workbooks and the long
table cannot drift apart. Outputs are generated artifacts and are git-ignored.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data_cleaning.company import CompanyData

from . import records

LOG = logging.getLogger("pfpa.submission")

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(_PROJECT_ROOT, "outputs")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _format(ws) -> None:
    ws.freeze_panes = "B2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = max((len(str(x.value)) for x in col if x.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), 40)


def write_submission(companies: list[CompanyData],
                     filename: str | None = None) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if filename is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"submission_{stamp}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)

    # Both frames come from dashboard.records, the single source of truth for
    # every published credit field. The Asset sheet is guaranteed to carry
    # exactly records.ASSET_SCHEMA, in order.
    readme = records.readme_frame(companies)
    asset = records.asset_frame(companies)
    validation = records.validation_frame(companies)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # README first: a reader opening the file lands on the provenance and
        # the conventions before the numbers.
        readme.to_excel(writer, sheet_name="README", index=False)
        asset.to_excel(writer, sheet_name="Asset", index=False)
        validation.to_excel(writer, sheet_name="validation", index=False)
        _format(writer.book["README"])
        _format(writer.book["Asset"])
        _format(writer.book["validation"])
        # The README is prose; give it room rather than content-fitting.
        ws = writer.book["README"]
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 110
        for row in ws.iter_rows(min_row=2):
            row[1].alignment = row[1].alignment.copy(wrap_text=True, vertical="top")

    LOG.info("submission workbook -> %s", os.path.relpath(path, _PROJECT_ROOT))
    return path
