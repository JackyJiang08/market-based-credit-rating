"""Grid mechanism tests against a SYNTHETIC workbook.

The licensed conversion grids never leave local/, so on any machine without
them (every CI cell) the loader, the grid lookup, the floor logic and the
letter mapping would otherwise go untested -- the coverage gate caught
exactly that. These tests exercise the mechanism with a tiny synthetic
workbook whose numbers are made up here; no licensed value appears.
"""

from __future__ import annotations

import numpy as np
import pytest
from creditrating.diagnostics import uncertainty as bs
from creditrating.model import conversion
from creditrating.tables import validation
from creditrating.tables.loader import load_tables
from openpyxl import Workbook


@pytest.fixture(scope="module")
def synthetic_tables(tmp_path_factory):
    """A 4x3 grid workbook in the loader's expected layout."""
    wb = Workbook()
    ccm_axis = [0.5, 1.0, 2.0, 4.0]
    mu_axis = [2.0, 10.0, 50.0]
    # Simple, monotone-ish fake PDs; the 2bp floor appears on purpose.
    ttc = [
        [0.0002, 0.0002, 0.0002],
        [0.0010, 0.0008, 0.0005],
        [0.0100, 0.0080, 0.0050],
        [0.1000, 0.0800, 0.0500],
    ]
    for sheet in ("TTC", "PIT"):
        ws = wb.create_sheet(sheet)
        ws.cell(row=2, column=1, value=sheet)  # header junk row (ignored)
        for j, mu in enumerate(mu_axis):
            ws.cell(row=2, column=2 + j, value=mu)
        for i, ccm in enumerate(ccm_axis):
            ws.cell(row=3 + i, column=1, value=ccm)
            for j in range(len(mu_axis)):
                ws.cell(row=3 + i, column=2 + j, value=ttc[i][j])
    sp = wb.create_sheet("SP")
    for i, (label, thr) in enumerate(
        [("AAA", 0.0001), ("AA", 0.0005), ("A", 0.002), ("BBB", 0.01), ("BB", 0.05)], start=1
    ):
        sp.cell(row=i, column=1, value=label)
        sp.cell(row=i, column=2, value=thr)
    wb.remove(wb["Sheet"])
    path = tmp_path_factory.mktemp("grids") / "synthetic.xlsx"
    wb.save(path)
    load_tables.cache_clear()
    tables = load_tables(str(path))
    load_tables.cache_clear()
    return tables


def test_loader_parses_axes_grid_and_thresholds(synthetic_tables):
    t = synthetic_tables
    assert list(t.ccm_axis[np.isfinite(t.ccm_axis)]) == [0.5, 1.0, 2.0, 4.0]
    assert list(t.mu_axis[np.isfinite(t.mu_axis)]) == [2.0, 10.0, 50.0]
    assert t.sp_labels == ["AAA", "AA", "A", "BBB", "BB"]
    assert validation.validate(t) == []


def test_interior_lookup_interpolates_and_reports_grid_basis(synthetic_tables):
    r = conversion.ttc_pd(synthetic_tables, ccm=1.5, mu=6.0)
    assert r.basis is conversion.RatingBasis.GRID_INTERIOR
    assert not r.off_grid
    assert 0.0005 <= r.value <= 0.01, "between the surrounding cells"


def test_floor_cells_are_flagged_floor_determined(synthetic_tables):
    r = conversion.ttc_pd(synthetic_tables, ccm=0.5, mu=2.0)
    assert r.at_floor, "the grid's smallest expressible value is a floor, not a measurement"


def test_off_grid_points_use_the_analytical_route(synthetic_tables):
    r = conversion.ttc_pd(synthetic_tables, ccm=100.0, mu=500.0, pit_pd=0.02)
    assert r.basis is conversion.RatingBasis.ANALYTICAL


def test_letters_map_through_thresholds(synthetic_tables):
    assert conversion.sp_rating(synthetic_tables, 0.00005) == "AAA"
    assert (
        conversion.sp_rating(synthetic_tables, 0.004) == "A"
    )  # worst grade whose threshold <= PD
    assert conversion.sp_rating(synthetic_tables, 0.9) == "BB"


def test_validation_flags_broken_tables(synthetic_tables):
    import dataclasses

    broken = dataclasses.replace(
        synthetic_tables,
        ccm_axis=np.array([2.0, 1.0]),
        ttc_grid=np.array([[1.5], [0.5]]),
        sp_thresholds=np.array([0.5, 0.1]),
    )
    problems = validation.validate(broken)
    assert any("ascending" in p for p in problems)
    assert any("outside [0, 1]" in p for p in problems)


def test_bootstrap_letter_interval_uses_the_synthetic_scale(synthetic_tables):
    rng = np.random.default_rng(7)
    u = rng.normal(0.0004, 0.02, 400)
    b = bs.run("SYN", u, 2.0e11, 8.0e10, synthetic_tables, n_replicates=60, seed=11)
    best, worst, width = b.notch_interval()
    assert best in synthetic_tables.sp_labels and worst in synthetic_tables.sp_labels
    assert width >= 1
